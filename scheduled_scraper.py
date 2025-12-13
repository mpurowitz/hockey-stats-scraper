#!/usr/bin/env python3
"""
Scheduled scraper that runs weekly to update hockey stats data.
Designed to run as a Render cron job.
"""

import os
import sys
import json
from datetime import datetime
import time
import importlib.util

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Import the scraper (handle hyphens in filename)
scraper_file = 'enhanced_scraper_2025-2026.py'
spec = importlib.util.spec_from_file_location("scraper_module", scraper_file)
scraper_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(scraper_module)
EliteProspectsScraper = scraper_module.EliteProspectsScraper

# Configuration
DATA_DIR = os.environ.get('DATA_DIR', '/data')
LEAGUES_TO_SCRAPE = [
    {
        'name': 'NA3HL',
        'url': 'https://www.eliteprospects.com/league/na3hl',
        'max_teams': None  # Scrape all teams
    },
    {
        'name': 'USPHL Premier',
        'url': 'https://www.eliteprospects.com/league/usphl-premier',
        'max_teams': None
    },
    {
        'name': 'USPHL Elite',
        'url': 'https://www.eliteprospects.com/league/usphl-elite',
        'max_teams': None
    },
    {
        'name': 'EHL',
        'url': 'https://www.eliteprospects.com/league/ehl',
        'max_teams': None
    },
    {
        'name': 'EHLP',
        'url': 'https://www.eliteprospects.com/league/ehlp',
        'max_teams': None
    },
    {
        'name': 'NCDC',
        'url': 'https://www.eliteprospects.com/league/ncdc',
        'max_teams': None
    },
    {
        'name': 'NAHL',
        'url': 'https://www.eliteprospects.com/league/nahl',
        'max_teams': None
    }
]

SEASON = '2025-2026'
DELAY = 3  # Seconds between requests

def save_data_to_files(scraped_data, timestamp):
    """Save scraped data to JSON and Excel files"""
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # Save as JSON
    json_filename = os.path.join(DATA_DIR, f'scraped_data_{timestamp}.json')
    with open(json_filename, 'w') as f:
        json.dump(scraped_data, f, indent=2)
    print(f"✅ Saved JSON: {json_filename}")
    
    # Save as Excel (using openpyxl)
    try:
        import openpyxl
        from openpyxl import Workbook
        
        excel_filename = os.path.join(DATA_DIR, f'hockey_stats_{timestamp}.xlsx')
        wb = Workbook()
        
        # Create a sheet for each league
        for league_name, teams in scraped_data.items():
            # Sanitize sheet name
            sheet_name = league_name.replace('/', '-')[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Collect all players from all teams
            all_players = []
            for team in teams:
                if 'players' in team and team['players']:
                    for player in team['players']:
                        player_row = {
                            'Name': player.get('name', ''),
                            'Jersey': player.get('jersey', ''),
                            'Position': player.get('position', ''),
                            'Shoots': player.get('shoots', ''),
                            'Age': player.get('age', ''),
                            'Birth Year': player.get('birthYear', ''),
                            'Height': player.get('height', ''),
                            'Weight': player.get('weight', ''),
                            'Hometown': player.get('hometown', ''),
                            'GP': player.get('games', 0),
                            'G': player.get('goals', 0),
                            'A': player.get('assists', 0),
                            'P': player.get('points', 0),
                            'PPG': player.get('ppg', 0.0),
                            'PIM': player.get('pim', 0),
                            'Team': team.get('name', ''),
                            'League': league_name,
                            'Season': player.get('season', SEASON),
                            'Profile URL': player.get('profile_url', '')
                        }
                        all_players.append(player_row)
            
            if all_players:
                # Write headers
                headers = list(all_players[0].keys())
                ws.append(headers)
                
                # Write data
                for player in all_players:
                    ws.append(list(player.values()))
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(excel_filename)
        print(f"✅ Saved Excel: {excel_filename}")
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")

def main():
    """Main scheduled scraping function"""
    print("\n" + "="*70)
    print(f"🏒 SCHEDULED SCRAPER STARTED")
    print(f"{'='*70}")
    print(f"⏰ Time: {datetime.now().isoformat()}")
    print(f"📁 Data directory: {DATA_DIR}")
    print(f"🏒 Leagues to scrape: {len(LEAGUES_TO_SCRAPE)}")
    print(f"{'='*70}\n")
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    scraped_data = {}
    scraper = None
    
    try:
        # Create scraper instance
        scraper = EliteProspectsScraper(
            headless=True,
            delay=DELAY,
            max_teams=None,  # Scrape all teams
            batch_size=5
        )
        
        # Scrape each league
        for i, league in enumerate(LEAGUES_TO_SCRAPE, 1):
            print(f"\n{'='*70}")
            print(f"🏒 LEAGUE {i}/{len(LEAGUES_TO_SCRAPE)}: {league['name']}")
            print(f"{'='*70}\n")
            
            try:
                # Get teams from league
                league_url = f"{league['url']}/{SEASON}"
                teams = scraper.get_league_teams(league_url, SEASON)
                
                if not teams:
                    print(f"⚠️ No teams found for {league['name']}")
                    continue
                
                print(f"📋 Found {len(teams)} teams in {league['name']}")
                
                # Scrape teams
                scraped_teams = scraper.scrape_multiple_teams(teams, SEASON)
                
                if scraped_teams:
                    scraped_data[league['name']] = scraped_teams
                    total_players = sum(len(team.get('players', [])) for team in scraped_teams)
                    print(f"\n✅ {league['name']} COMPLETE: {len(scraped_teams)} teams, {total_players} players")
                else:
                    print(f"⚠️ No data scraped for {league['name']}")
                
            except Exception as e:
                print(f"❌ Error scraping {league['name']}: {e}")
                continue
            
            # Brief pause between leagues
            time.sleep(5)
        
        # Save all data
        if scraped_data:
            print(f"\n{'='*70}")
            print(f"💾 SAVING DATA")
            print(f"{'='*70}\n")
            save_data_to_files(scraped_data, timestamp)
            
            # Summary
            total_leagues = len(scraped_data)
            total_teams = sum(len(teams) for teams in scraped_data.values())
            total_players = sum(
                sum(len(team.get('players', [])) for team in teams)
                for teams in scraped_data.values()
            )
            
            print(f"\n{'='*70}")
            print(f"✅ SCRAPING COMPLETE")
            print(f"{'='*70}")
            print(f"📊 Leagues scraped: {total_leagues}")
            print(f"🏒 Teams scraped: {total_teams}")
            print(f"👤 Players scraped: {total_players}")
            print(f"⏰ Timestamp: {timestamp}")
            print(f"{'='*70}\n")
        else:
            print("\n❌ No data was scraped")
            sys.exit(1)
        
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    finally:
        # Cleanup
        if scraper:
            try:
                scraper.close()
                print("🧹 Scraper cleaned up")
            except:
                pass

if __name__ == '__main__':
    main()
