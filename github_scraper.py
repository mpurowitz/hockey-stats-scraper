#!/usr/bin/env python3
"""
GitHub Actions scraper - runs weekly and saves data to /data folder.
Data is committed to the repo and served via GitHub Pages.
"""

import os
import sys
import json
from datetime import datetime
import importlib.util

# Import the scraper (handle hyphens in filename)
scraper_file = 'enhanced_scraper_2025-2026.py'
spec = importlib.util.spec_from_file_location("scraper_module", scraper_file)
scraper_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(scraper_module)
EliteProspectsScraper = scraper_module.EliteProspectsScraper

# Configuration
DATA_DIR = 'data'
LEAGUES_TO_SCRAPE = [
    {
        'name': 'NA3HL',
        'url': 'https://www.eliteprospects.com/league/na3hl',
        'max_teams': None
    },
    {
        'name': 'USPHL Premier',
        'url': 'https://www.eliteprospects.com/league/usphl-premier',
        'max_teams': None
    },
    {
        'name': 'USPHL Elite',
        'url': 'https://www.eliteprospects.com/league/usphl-elite',
        'max_teams': None
    },
    {
        'name': 'EHL',
        'url': 'https://www.eliteprospects.com/league/ehl',
        'max_teams': None
    },
    {
        'name': 'EHLP',
        'url': 'https://www.eliteprospects.com/league/ehlp',
        'max_teams': None
    },
    {
        'name': 'NCDC',
        'url': 'https://www.eliteprospects.com/league/ncdc',
        'max_teams': None
    },
    {
        'name': 'NAHL',
        'url': 'https://www.eliteprospects.com/league/nahl',
        'max_teams': None
    }
]

SEASON = '2025-2026'
DELAY = 3

def save_data(scraped_data, timestamp):
    """Save scraped data to JSON files"""
    os.makedirs(DATA_DIR, exist_ok=True)
    
    # Save with timestamp
    dated_file = os.path.join(DATA_DIR, f'{timestamp}.json')
    with open(dated_file, 'w') as f:
        json.dump(scraped_data, f, indent=2)
    print(f"✅ Saved: {dated_file}")
    
    # Save as 'latest.json' for dashboard
    latest_file = os.path.join(DATA_DIR, 'latest.json')
    with open(latest_file, 'w') as f:
        json.dump(scraped_data, f, indent=2)
    print(f"✅ Saved: {latest_file}")
    
    # Create index of all data files
    data_files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.json') and f != 'latest.json' and f != 'index.json'])
    index = {
        'last_updated': timestamp,
        'total_scrapes': len(data_files),
        'files': data_files
    }
    
    index_file = os.path.join(DATA_DIR, 'index.json')
    with open(index_file, 'w') as f:
        json.dump(index, f, indent=2)
    print(f"✅ Updated index: {index_file}")

def create_excel(scraped_data, timestamp):
    """Create Excel file"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        excel_file = os.path.join(DATA_DIR, f'{timestamp}.xlsx')
        wb = Workbook()
        
        for league_name, teams in scraped_data.items():
            sheet_name = league_name.replace('/', '-')[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            all_players = []
            for team in teams:
                if 'players' in team and team['players']:
                    for player in team['players']:
                        all_players.append({
                            'Name': player.get('name', ''),
                            'Jersey': player.get('jersey', ''),
                            'Position': player.get('position', ''),
                            'Shoots': player.get('shoots', ''),
                            'Age': player.get('age', ''),
                            'Birth Year': player.get('birthYear', ''),
                            'Height': player.get('height', ''),
                            'Weight': player.get('weight', ''),
                            'Hometown': player.get('hometown', ''),
                            'GP': player.get('games', 0),
                            'G': player.get('goals', 0),
                            'A': player.get('assists', 0),
                            'P': player.get('points', 0),
                            'PPG': player.get('ppg', 0.0),
                            'PIM': player.get('pim', 0),
                            'Team': team.get('name', ''),
                            'League': league_name,
                            'Season': player.get('season', SEASON),
                            'Profile URL': player.get('profile_url', '')
                        })
            
            if all_players:
                headers = list(all_players[0].keys())
                ws.append(headers)
                for player in all_players:
                    ws.append(list(player.values()))
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(excel_file)
        print(f"✅ Created Excel: {excel_file}")
        
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")

def main():
    print("\n" + "="*70)
    print("🏒 GITHUB ACTIONS SCRAPER")
    print("="*70)
    print(f"⏰ Time: {datetime.now().isoformat()}")
    print(f"📁 Data directory: {DATA_DIR}")
    print(f"🏒 Leagues: {len(LEAGUES_TO_SCRAPE)}")
    print("="*70 + "\n")
    
    timestamp = datetime.now().strftime('%Y-%m-%d')
    scraped_data = {}
    scraper = None
    
    try:
        scraper = EliteProspectsScraper(
            headless=True,
            delay=DELAY,
            max_teams=None,
            batch_size=5
        )
        
        for i, league in enumerate(LEAGUES_TO_SCRAPE, 1):
            print(f"\n{'='*70}")
            print(f"🏒 LEAGUE {i}/{len(LEAGUES_TO_SCRAPE)}: {league['name']}")
            print(f"{'='*70}\n")
            
            try:
                league_url = f"{league['url']}/{SEASON}"
                teams = scraper.get_league_teams(league_url, SEASON)
                
                if not teams:
                    print(f"⚠️ No teams found for {league['name']}")
                    continue
                
                print(f"📋 Found {len(teams)} teams in {league['name']}")
                
                if league.get('max_teams'):
                    teams = teams[:league['max_teams']]
                
                scraped_teams = scraper.scrape_multiple_teams(teams, SEASON)
                
                if scraped_teams:
                    scraped_data[league['name']] = scraped_teams
                    total_players = sum(len(team.get('players', [])) for team in scraped_teams)
                    print(f"\n✅ {league['name']} COMPLETE: {len(scraped_teams)} teams, {total_players} players")
                
            except Exception as e:
                print(f"❌ Error scraping {league['name']}: {e}")
                continue
        
        if scraped_data:
            print(f"\n{'='*70}")
            print("💾 SAVING DATA")
            print(f"{'='*70}\n")
            
            save_data(scraped_data, timestamp)
            create_excel(scraped_data, timestamp)
            
            total_leagues = len(scraped_data)
            total_teams = sum(len(teams) for teams in scraped_data.values())
            total_players = sum(
                sum(len(team.get('players', [])) for team in teams)
                for teams in scraped_data.values()
            )
            
            print(f"\n{'='*70}")
            print("✅ SCRAPING COMPLETE")
            print(f"{'='*70}")
            print(f"📊 Leagues: {total_leagues}")
            print(f"🏒 Teams: {total_teams}")
            print(f"👤 Players: {total_players}")
            print(f"⏰ Date: {timestamp}")
            print(f"{'='*70}\n")
        else:
            print("\n❌ No data scraped")
            sys.exit(1)
        
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    finally:
        if scraper:
            try:
                scraper.close()
            except:
                pass

if __name__ == '__main__':
    main()
